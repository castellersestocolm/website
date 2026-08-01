import pytest
from django.test import TestCase
from django.utils import timezone
from djmoney.money import Money
from openpyxl.reader.excel import load_workbook

from comunicat.enums import Module
from conftest import NumOperationsMixin
from membership.api.export import export_memberships
from membership.enums import MembershipStatus
from membership.tests.factories import (
    MembershipFactory,
    MembershipModuleFactory,
    MembershipUserFactory,
)
from user.enums import FamilyMemberRole, FamilyMemberStatus
from user.tests.factories import FamilyFactory, FamilyMemberFactory, UserFactory


@pytest.mark.django_db
class TestExportMemberships(NumOperationsMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        cls.user_member_1_obj = UserFactory(
            firstname="firstname-1",
            lastname="lastname-1",
            email="user-member-1@domain-test.org",
        )
        cls.user_member_2_obj = UserFactory(
            firstname="firstname-2",
            lastname="lastname-2",
            email="user-member-2@domain-test.org",
        )
        cls.user_member_3_obj = UserFactory(
            firstname="firstname-3",
            lastname="lastname-3",
            email="user-member-3@domain-test.org",
        )
        cls.user_member_4_obj = UserFactory(
            firstname="firstname-4",
            lastname="lastname-4",
            email="user-member-4@domain-test.org",
        )

        cls.membership_1_user_1_obj = MembershipFactory(
            status=MembershipStatus.ACTIVE,
            date_from=timezone.localdate() - timezone.timedelta(days=365),
            date_to=timezone.localdate() - timezone.timedelta(days=1),
        )
        cls.membership_2_user_1_obj = MembershipFactory(
            status=MembershipStatus.PROCESSING,
            date_from=timezone.localdate(),
            date_to=timezone.localdate() + timezone.timedelta(days=365 - 1),
        )
        cls.membership_3_user_2_obj = MembershipFactory(
            status=MembershipStatus.ACTIVE,
            date_from=timezone.localdate(),
            date_to=timezone.localdate() + timezone.timedelta(days=365 - 1),
        )

        cls.membership_module_1_user_1_obj = MembershipModuleFactory(
            membership=cls.membership_1_user_1_obj,
            status=MembershipStatus.ACTIVE,
            amount=Money(150, "SEK"),
            module=Module.ORG,
        )
        cls.membership_module_2_user_2_obj = MembershipModuleFactory(
            membership=cls.membership_2_user_1_obj,
            status=MembershipStatus.PROCESSING,
            amount=Money(150, "SEK"),
            module=Module.ORG,
        )
        cls.membership_module_3_user_3_obj = MembershipModuleFactory(
            membership=cls.membership_3_user_2_obj,
            status=MembershipStatus.ACTIVE,
            amount=Money(250, "SEK"),
            module=Module.ORG,
        )
        cls.membership_module_4_user_3_obj = MembershipModuleFactory(
            membership=cls.membership_3_user_2_obj,
            status=MembershipStatus.ACTIVE,
            amount=Money(450, "SEK"),
            module=Module.TOWERS,
        )

        cls.family_1_obj = FamilyFactory()

        cls.family_1_user_2_obj = FamilyMemberFactory(
            user=cls.user_member_2_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_1_user_3_obj = FamilyMemberFactory(
            user=cls.user_member_3_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_1_user_4_obj = FamilyMemberFactory(
            user=cls.user_member_4_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MEMBER,
            status=FamilyMemberStatus.ACTIVE,
        )

        cls.membership_user_1_user_1_obj = MembershipUserFactory(
            user=cls.user_member_1_obj,
            membership=cls.membership_1_user_1_obj,
            family=None,
        )
        cls.membership_user_2_user_1_obj = MembershipUserFactory(
            user=cls.user_member_1_obj,
            membership=cls.membership_2_user_1_obj,
            family=None,
        )
        cls.membership_user_3_user_2_obj = MembershipUserFactory(
            user=cls.user_member_2_obj,
            membership=cls.membership_3_user_2_obj,
            family=cls.family_1_obj,
        )
        cls.membership_user_4_user_3_obj = MembershipUserFactory(
            user=cls.user_member_3_obj,
            membership=cls.membership_3_user_2_obj,
            family=cls.family_1_obj,
        )
        cls.membership_user_5_user_4_obj = MembershipUserFactory(
            user=cls.user_member_4_obj,
            membership=cls.membership_3_user_2_obj,
            family=cls.family_1_obj,
        )

    def test_export_memberships(self, *args, **kwargs):
        with self.assertNumOperations(num=0, num_selects=12):
            memberships_file = export_memberships(module=Module.ORG)

        self.assertIsNotNone(memberships_file)

        wb = load_workbook(filename=memberships_file)

        self.assertEqual(wb.sheetnames, ["Registrations"])

        ws = wb.worksheets[0]
        ws_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        self.assertEqual(len(ws_rows), 6)
        self.assertEqual(
            ws_rows[0],
            [
                "First name",
                "Last name",
                "Email",
                "Phone",
                "Family",
                "Membership",
                "Date from",
                "Date to",
                "Status",
                "Price",
            ],
        )
        self.assertEqual(
            ws_rows[1],
            [
                "firstname-1",
                "lastname-1",
                "user-member-1@domain-test.org",
                self.user_member_1_obj.phone,
                None,
                "Name ORG",
                self.membership_1_user_1_obj.date_from.strftime("%Y-%m-%d"),
                self.membership_1_user_1_obj.date_to.strftime("%Y-%m-%d"),
                "ACTIVE",
                "150",
            ],
        )
        self.assertEqual(
            ws_rows[2],
            [
                "firstname-1",
                "lastname-1",
                "user-member-1@domain-test.org",
                self.user_member_1_obj.phone,
                None,
                "Name ORG",
                self.membership_2_user_1_obj.date_from.strftime("%Y-%m-%d"),
                self.membership_2_user_1_obj.date_to.strftime("%Y-%m-%d"),
                "PROCESSING",
                "150",
            ],
        )
        self.assertEqual(
            ws_rows[3],
            [
                "firstname-2",
                "lastname-2",
                "user-member-2@domain-test.org",
                self.user_member_2_obj.phone,
                "lastname-2-lastname-3",
                "Name ORG\nName TOWERS",
                self.membership_3_user_2_obj.date_from.strftime("%Y-%m-%d"),
                self.membership_3_user_2_obj.date_to.strftime("%Y-%m-%d"),
                "ACTIVE",
                "250",
            ],
        )
        self.assertEqual(
            ws_rows[4],
            [
                "firstname-3",
                "lastname-3",
                "user-member-3@domain-test.org",
                self.user_member_3_obj.phone,
                "lastname-2-lastname-3",
                "Name ORG\nName TOWERS",
                self.membership_3_user_2_obj.date_from.strftime("%Y-%m-%d"),
                self.membership_3_user_2_obj.date_to.strftime("%Y-%m-%d"),
                "ACTIVE",
                "250",
            ],
        )
        self.assertEqual(
            ws_rows[5],
            [
                "firstname-4",
                "lastname-4",
                "user-member-4@domain-test.org",
                self.user_member_4_obj.phone,
                "lastname-2-lastname-3",
                "Name ORG\nName TOWERS",
                self.membership_3_user_2_obj.date_from.strftime("%Y-%m-%d"),
                self.membership_3_user_2_obj.date_to.strftime("%Y-%m-%d"),
                "ACTIVE",
                "250",
            ],
        )
