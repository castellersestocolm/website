import pytest
from django.test import TestCase
from django.utils import timezone
from djmoney.money import Money
from openpyxl.reader.excel import load_workbook

from activity.api.export import export_program_course
from activity.enums import ProgramCourseRegistrationStatus, ProgramType
from activity.tests.factories import (
    ProgramCourseFactory,
    ProgramCoursePriceFactory,
    ProgramCourseRegistrationFactory,
    ProgramFactory,
)
from comunicat.enums import Module
from conftest import NumOperationsMixin
from payment.tests.factories import EntityFactory
from user.enums import FamilyMemberRole, FamilyMemberStatus
from user.tests.factories import FamilyFactory, FamilyMemberFactory, UserFactory


@pytest.mark.django_db
class TestExportProgramCourse(NumOperationsMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        cls.program_1_obj = ProgramFactory(module=Module.ORG, type=ProgramType.ESPLAI)
        cls.program_2_obj = ProgramFactory(module=Module.ORG, type=ProgramType.TEACHING)
        cls.program_3_obj = ProgramFactory(
            module=Module.TOWERS, type=ProgramType.GENERAL
        )

        cls.program_course_1_obj = ProgramCourseFactory(
            program=cls.program_1_obj,
        )
        cls.program_course_2_obj = ProgramCourseFactory(
            program=cls.program_2_obj,
        )
        cls.program_course_3_obj = ProgramCourseFactory(
            program=cls.program_3_obj,
        )

        cls.program_course_price_1_obj = ProgramCoursePriceFactory(
            course=cls.program_course_1_obj, amount=Money(100, "SEK")
        )
        cls.program_course_price_2_obj = ProgramCoursePriceFactory(
            course=cls.program_course_2_obj, amount=Money(500, "SEK")
        )
        cls.program_course_price_3_obj = ProgramCoursePriceFactory(
            course=cls.program_course_3_obj, amount=Money(250, "SEK")
        )

        cls.user_member_1_obj = UserFactory(
            firstname="firstname-1",
            lastname="lastname-1",
            email="user-member-1@domain-test.org",
        )
        cls.user_member_2_obj = UserFactory(
            firstname="firstname-2",
            lastname="lastname-2",
            email="user-member-1+user-member-2@domain-test.org",
            birthday=timezone.localdate().replace(year=timezone.localdate().year - 8),
        )
        cls.user_member_3_obj = UserFactory(
            firstname="firstname-3",
            lastname="lastname-3",
            email="user-member-1+user-member-3@domain-test.org",
            birthday=timezone.localdate().replace(year=timezone.localdate().year - 12),
        )
        cls.user_member_4_obj = UserFactory(
            firstname="firstname-4",
            lastname="lastname-4",
            email="user-member-4@domain-test.org",
        )
        cls.user_member_5_obj = UserFactory(
            firstname="firstname-5",
            lastname="lastname-5",
            email="user-member-5@domain-test.org",
        )
        cls.user_member_6_obj = UserFactory(
            firstname="firstname-6",
            lastname="lastname-6",
            email="user-member-6+user-member-4@domain-test.org",
            birthday=timezone.localdate().replace(year=timezone.localdate().year - 10),
        )

        cls.entity_member_1_obj = EntityFactory(user=cls.user_member_1_obj)
        cls.entity_member_2_obj = EntityFactory(user=cls.user_member_2_obj)
        cls.entity_member_3_obj = EntityFactory(user=cls.user_member_3_obj)
        cls.entity_member_4_obj = EntityFactory(user=cls.user_member_4_obj)
        cls.entity_member_5_obj = EntityFactory(user=cls.user_member_5_obj)
        cls.entity_member_6_obj = EntityFactory(user=cls.user_member_6_obj)

        cls.family_1_obj = FamilyFactory()
        cls.family_2_obj = FamilyFactory()

        cls.family_1_user_1_obj = FamilyMemberFactory(
            user=cls.user_member_1_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_1_user_2_obj = FamilyMemberFactory(
            user=cls.user_member_2_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MEMBER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_1_user_3_obj = FamilyMemberFactory(
            user=cls.user_member_3_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MEMBER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_2_user_4_obj = FamilyMemberFactory(
            user=cls.user_member_4_obj,
            family=cls.family_2_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_2_user_5_obj = FamilyMemberFactory(
            user=cls.user_member_5_obj,
            family=cls.family_2_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_2_user_6_obj = FamilyMemberFactory(
            user=cls.user_member_6_obj,
            family=cls.family_2_obj,
            role=FamilyMemberRole.MEMBER,
            status=FamilyMemberStatus.ACTIVE,
        )

        cls.program_course_1_user_2_obj = ProgramCourseRegistrationFactory(
            course=cls.program_course_1_obj,
            entity=cls.entity_member_2_obj,
            status=ProgramCourseRegistrationStatus.ACTIVE,
            price=cls.program_course_price_1_obj,
        )
        cls.program_course_1_user_3_obj = ProgramCourseRegistrationFactory(
            course=cls.program_course_1_obj,
            entity=cls.entity_member_3_obj,
            status=ProgramCourseRegistrationStatus.ACTIVE,
            price=cls.program_course_price_1_obj,
        )
        cls.program_course_1_user_6_obj = ProgramCourseRegistrationFactory(
            course=cls.program_course_1_obj,
            entity=cls.entity_member_6_obj,
            status=ProgramCourseRegistrationStatus.ACTIVE,
            price=cls.program_course_price_1_obj,
        )
        cls.program_course_2_user_6_obj = ProgramCourseRegistrationFactory(
            course=cls.program_course_2_obj,
            entity=cls.entity_member_6_obj,
            status=ProgramCourseRegistrationStatus.REQUESTED,
            price=cls.program_course_price_2_obj,
        )
        cls.program_course_3_user_6_obj = ProgramCourseRegistrationFactory(
            course=cls.program_course_3_obj,
            entity=cls.entity_member_6_obj,
            status=ProgramCourseRegistrationStatus.ACTIVE,
            price=cls.program_course_price_3_obj,
        )

    # TODO: Test attendance sheet
    def test_export_program_course__program_course_1(self, *args, **kwargs):
        with self.assertNumOperations(num=0, num_selects=9):
            program_course_file = export_program_course(
                course_id=self.program_course_1_obj.id
            )

        self.assertIsNotNone(program_course_file)

        wb = load_workbook(filename=program_course_file)

        self.assertEqual(wb.sheetnames, ["Inscripcions", "Assistència"])

        ws = wb.worksheets[0]
        ws_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        self.assertEqual(len(ws_rows), 4)
        self.assertEqual(
            ws_rows[0],
            [
                "Nom",
                "Cognom",
                "Edat",
                "Família",
                "Pares",
                "Correus electrònics",
                "Telèfons",
                "Estat",
                "Preu",
            ],
        )
        self.assertEqual(
            ws_rows[1],
            [
                "firstname-2",
                "lastname-2",
                8,
                "lastname-1",
                "firstname-1 lastname-1",
                "user-member-1@domain-test.org",
                self.user_member_1_obj.phone,
                "Actiu",
                100,
            ],
        )
        self.assertEqual(
            ws_rows[2],
            [
                "firstname-3",
                "lastname-3",
                12,
                "lastname-1",
                "firstname-1 lastname-1",
                "user-member-1@domain-test.org",
                self.user_member_1_obj.phone,
                "Actiu",
                100,
            ],
        )
        self.assertEqual(
            ws_rows[3],
            [
                "firstname-6",
                "lastname-6",
                10,
                "lastname-4-lastname-5",
                "firstname-4 lastname-4\nfirstname-5 lastname-5",
                "user-member-4@domain-test.org\nuser-member-5@domain-test.org",
                f"{self.user_member_4_obj.phone}\n{self.user_member_5_obj.phone}",
                "Actiu",
                100,
            ],
        )

    def test_export_program_course__program_course_2(self, *args, **kwargs):
        with self.assertNumOperations(num=0, num_selects=7):
            program_course_file = export_program_course(
                course_id=self.program_course_2_obj.id
            )

        self.assertIsNotNone(program_course_file)

        wb = load_workbook(filename=program_course_file)

        self.assertEqual(wb.sheetnames, ["Inscripcions", "Assistència"])

        ws = wb.worksheets[0]
        ws_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        self.assertEqual(len(ws_rows), 2)
        self.assertEqual(
            ws_rows[0],
            [
                "Nom",
                "Cognom",
                "Edat",
                "Família",
                "Pares",
                "Correus electrònics",
                "Telèfons",
                "Estat",
                "Preu",
            ],
        )
        self.assertEqual(
            ws_rows[1],
            [
                "firstname-6",
                "lastname-6",
                10,
                "lastname-4-lastname-5",
                "firstname-4 lastname-4\nfirstname-5 lastname-5",
                "user-member-4@domain-test.org\nuser-member-5@domain-test.org",
                f"{self.user_member_4_obj.phone}\n{self.user_member_5_obj.phone}",
                "Sol·licitat",
                500,
            ],
        )

    def test_export_program_course__program_course_3(self, *args, **kwargs):
        with self.assertNumOperations(num=0, num_selects=7):
            program_course_file = export_program_course(
                course_id=self.program_course_3_obj.id
            )

        self.assertIsNotNone(program_course_file)

        wb = load_workbook(filename=program_course_file)

        self.assertEqual(wb.sheetnames, ["Registrations", "Attendance"])

        ws = wb.worksheets[0]
        ws_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        self.assertEqual(len(ws_rows), 2)
        self.assertEqual(
            ws_rows[0],
            [
                "First name",
                "Last name",
                "Age",
                "Family",
                "Parents",
                "Emails",
                "Phones",
                "Status",
                "Price",
            ],
        )
        self.assertEqual(
            ws_rows[1],
            [
                "firstname-6",
                "lastname-6",
                10,
                "lastname-4-lastname-5",
                "firstname-4 lastname-4\nfirstname-5 lastname-5",
                "user-member-4@domain-test.org\nuser-member-5@domain-test.org",
                f"{self.user_member_4_obj.phone}\n{self.user_member_5_obj.phone}",
                "Active",
                250,
            ],
        )
