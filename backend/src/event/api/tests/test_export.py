import pytest
from django.test import TestCase
from django.utils import timezone
from djmoney.money import Money
from openpyxl.reader.excel import load_workbook
from social_core.utils import first

from comunicat.enums import Module
from conftest import NumOperationsMixin
from event.api.export import export_event
from event.enums import EventType, RegistrationStatus
from event.tests.factories import EventFactory, EventPriceFactory, RegistrationFactory
from payment.tests.factories import EntityFactory
from user.enums import FamilyMemberRole, FamilyMemberStatus
from user.tests.factories import FamilyFactory, FamilyMemberFactory, UserFactory


@pytest.mark.django_db
class TestExportEvent(NumOperationsMixin, TestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        cls.event_1_obj = EventFactory(module=Module.ORG, type=EventType.GATHERING)
        cls.event_2_obj = EventFactory(module=Module.ORG, type=EventType.PERFORMANCE)

        cls.event_price_1_obj = EventPriceFactory(
            event=cls.event_1_obj, amount=Money(100, "SEK")
        )
        cls.event_price_2_obj = EventPriceFactory(
            event=cls.event_1_obj, amount=Money(500, "SEK")
        )

        cls.user_member_1_obj = UserFactory(
            firstname="firstname-1",
            lastname="lastname-1",
            email="user-member-1@domain-test.org",
        )
        cls.user_member_2_obj = UserFactory(
            firstname="firstname-2",
            lastname="lastname-2",
            email="user-member-1+user-member-2@domain-test.org",
            birthday=timezone.localdate().replace(year=timezone.localdate().year - 8),
        )
        cls.user_member_3_obj = UserFactory(
            firstname="firstname-3",
            lastname="lastname-3",
            email="user-member-1+user-member-3@domain-test.org",
            birthday=timezone.localdate().replace(year=timezone.localdate().year - 12),
        )
        cls.user_member_4_obj = UserFactory(
            firstname="firstname-4",
            lastname="lastname-4",
            email="user-member-4@domain-test.org",
        )
        cls.user_member_5_obj = UserFactory(
            firstname="firstname-5",
            lastname="lastname-5",
            email="user-member-5@domain-test.org",
        )

        cls.entity_member_1_obj = EntityFactory(user=cls.user_member_1_obj)
        cls.entity_member_2_obj = EntityFactory(user=cls.user_member_2_obj)
        cls.entity_member_3_obj = EntityFactory(user=cls.user_member_3_obj)
        cls.entity_member_4_obj = EntityFactory(user=cls.user_member_4_obj)
        cls.entity_member_5_obj = EntityFactory(user=cls.user_member_5_obj)

        cls.entity_external_1_obj = EntityFactory(
            firstname="firstname-6",
            lastname="lastname-6",
            email="external-1@domain-test.org",
        )
        cls.entity_external_2_obj = EntityFactory(
            firstname="firstname-7",
            lastname="lastname-7",
            email="external-2@domain-test.org",
        )

        cls.family_1_obj = FamilyFactory()
        cls.family_2_obj = FamilyFactory()

        cls.family_1_user_1_obj = FamilyMemberFactory(
            user=cls.user_member_1_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_1_user_2_obj = FamilyMemberFactory(
            user=cls.user_member_2_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MEMBER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_1_user_3_obj = FamilyMemberFactory(
            user=cls.user_member_3_obj,
            family=cls.family_1_obj,
            role=FamilyMemberRole.MEMBER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_2_user_4_obj = FamilyMemberFactory(
            user=cls.user_member_4_obj,
            family=cls.family_2_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )
        cls.family_2_user_5_obj = FamilyMemberFactory(
            user=cls.user_member_5_obj,
            family=cls.family_2_obj,
            role=FamilyMemberRole.MANAGER,
            status=FamilyMemberStatus.ACTIVE,
        )

        cls.registration_event_1_user_1_obj = RegistrationFactory(
            event=cls.event_1_obj,
            entity=cls.entity_member_1_obj,
            status=RegistrationStatus.ACTIVE,
            price=cls.event_price_1_obj,
        )
        cls.registration_event_1_user_2_obj = RegistrationFactory(
            event=cls.event_1_obj,
            entity=cls.entity_member_2_obj,
            status=RegistrationStatus.ACTIVE,
            price=cls.event_price_1_obj,
        )
        cls.registration_event_1_user_3_obj = RegistrationFactory(
            event=cls.event_1_obj,
            entity=cls.entity_member_3_obj,
            status=RegistrationStatus.CANCELLED,
            price=cls.event_price_1_obj,
        )
        cls.registration_event_1_external_1_obj = RegistrationFactory(
            event=cls.event_1_obj,
            entity=cls.entity_external_1_obj,
            status=RegistrationStatus.ACTIVE,
            price=cls.event_price_1_obj,
        )
        cls.registration_event_2_user_4_obj = RegistrationFactory(
            event=cls.event_2_obj,
            entity=cls.entity_member_4_obj,
            status=RegistrationStatus.ACTIVE,
            price=cls.event_price_2_obj,
        )
        cls.registration_event_2_user_5_obj = RegistrationFactory(
            event=cls.event_2_obj,
            entity=cls.entity_member_5_obj,
            status=RegistrationStatus.ACTIVE,
            price=cls.event_price_2_obj,
        )
        cls.registration_event_2_external_2_obj = RegistrationFactory(
            event=cls.event_2_obj,
            entity=cls.entity_external_2_obj,
            status=RegistrationStatus.REQUESTED,
            price=cls.event_price_2_obj,
        )

    def test_export_event__event_1(self, *args, **kwargs):
        with self.assertNumOperations(num=0, num_selects=6):
            event_file = export_event(event_id=self.event_1_obj.id)

        self.assertIsNotNone(event_file)

        wb = load_workbook(filename=event_file)

        self.assertEqual(wb.sheetnames, ["Inscripcions"])

        ws = wb.worksheets[0]
        ws_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        self.assertEqual(len(ws_rows), 5)
        self.assertEqual(
            ws_rows[0],
            [
                "Nom",
                "Cognom",
                "Edat",
                "Família",
                "Correu electrònic",
                "Telèfon",
                "Estat",
                "Preu",
            ],
        )
        self.assertEqual(
            ws_rows[1],
            [
                "firstname-1",
                "lastname-1",
                self.user_member_1_obj.age,
                "lastname-1",
                "user-member-1@domain-test.org",
                self.user_member_1_obj.phone,
                "Actiu",
                100,
            ],
        )
        self.assertEqual(
            ws_rows[2],
            [
                "firstname-2",
                "lastname-2",
                8,
                "lastname-1",
                "user-member-1+user-member-2@domain-test.org",
                self.user_member_2_obj.phone,
                "Actiu",
                100,
            ],
        )
        self.assertEqual(
            ws_rows[3],
            [
                "firstname-3",
                "lastname-3",
                12,
                "lastname-1",
                "user-member-1+user-member-3@domain-test.org",
                self.user_member_3_obj.phone,
                "Cancel·lat",
                100,
            ],
        )
        self.assertEqual(
            ws_rows[4],
            [
                "firstname-6",
                "lastname-6",
                self.entity_external_1_obj.age,
                "lastname-6",
                "external-1@domain-test.org",
                self.entity_external_1_obj.phone,
                "Actiu",
                100,
            ],
        )

    def test_export_event__event_2(self, *args, **kwargs):
        with self.assertNumOperations(num=0, num_selects=5):
            event_file = export_event(event_id=self.event_2_obj.id)

        self.assertIsNotNone(event_file)

        wb = load_workbook(filename=event_file)

        self.assertEqual(wb.sheetnames, ["Inscripcions"])

        ws = wb.worksheets[0]
        ws_rows = [[cell.value for cell in row] for row in ws.iter_rows()]

        self.assertEqual(len(ws_rows), 4)
        self.assertEqual(
            ws_rows[0],
            [
                "Nom",
                "Cognom",
                "Edat",
                "Família",
                "Correu electrònic",
                "Telèfon",
                "Estat",
                "Preu",
            ],
        )
        self.assertEqual(
            ws_rows[1],
            [
                "firstname-4",
                "lastname-4",
                self.user_member_4_obj.age,
                "lastname-4-lastname-5",
                "user-member-4@domain-test.org",
                self.user_member_4_obj.phone,
                "Actiu",
                500,
            ],
        )
        self.assertEqual(
            ws_rows[2],
            [
                "firstname-5",
                "lastname-5",
                self.user_member_5_obj.age,
                "lastname-4-lastname-5",
                "user-member-5@domain-test.org",
                self.user_member_5_obj.phone,
                "Actiu",
                500,
            ],
        )
        self.assertEqual(
            ws_rows[3],
            [
                "firstname-7",
                "lastname-7",
                self.entity_external_1_obj.age,
                "lastname-7",
                "external-2@domain-test.org",
                self.entity_external_2_obj.phone,
                "Sol·licitat",
                500,
            ],
        )
